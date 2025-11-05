import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from models.entities import Operator, Vulnerability
from typing import List


class ExportService:
    @staticmethod
    def export_to_excel(data, filename_prefix="report"):
        """Экспорт данных в Excel файл"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Отчет"

        # Заголовки
        headers = list(data[0].keys()) if data else []
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Данные
        for row, item in enumerate(data, 2):
            for col, key in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=item.get(key, ''))

        # Авто-ширина колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        filename = f"{filename_prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(filename)
        return filename

    def export_vulnerabilities_report(self, operators: List[Operator]) -> str:
        """Экспорт отчета по уязвимостям операторов"""
        data = []
        for op in operators:
            for vuln in op.assigned_vulnerabilities:
                data.append({
                    'Имя оператора': op.name,
                    'Email оператора': op.email,
                    'ID уязвимости': vuln.id,
                    'Название уязвимости': vuln.title,
                    'Уровень риска': vuln.severity,
                    'Статус': vuln.status,
                    'CVSS Score': vuln.cvss_score,
                    'Категория': vuln.category,
                    'Дата назначения': datetime.now().strftime('%Y-%m-%d'),
                    'Метрика оператора': f"{op.current_metric}%"
                })

        return self.export_to_excel(data, "уязвимости_операторов")

    def export_operator_vulnerabilities(self, operators: List[Operator]) -> str:
        """Экспорт уязвимостей по каждому оператору"""
        data = []

        for op in operators:
            # Добавляем заголовок оператора
            data.append({
                'Оператор': f"ОПЕРАТОР: {op.name}",
                'Email': op.email,
                'Метрика': f"{op.current_metric}%",
                'Уровень опыта': f"{op.experience_level}%",
                'Нагрузка': f"{op.calculate_workload()}%",
                'Кол-во уязвимостей': len(op.assigned_vulnerabilities),
                'Статус': '',
                'CVSS': '',
                'Категория': ''
            })

            # Добавляем уязвимости оператора
            for vuln in op.assigned_vulnerabilities:
                data.append({
                    'Оператор': vuln.title,
                    'Email': '',
                    'Метрика': '',
                    'Уровень опыта': '',
                    'Нагрузка': '',
                    'Кол-во уязвимостей': '',
                    'Статус': vuln.status,
                    'CVSS': vuln.cvss_score,
                    'Категория': vuln.category
                })

            # Добавляем пустую строку между операторами
            data.append({})

        return self.export_to_excel(data, "отчет_операторов_уязвимостей")

    def export_single_operator_vulnerabilities(self, operator: Operator) -> str:
        """Экспорт уязвимостей для одного оператора"""
        data = []

        # Заголовок оператора
        data.append({
            'Оператор': operator.name,
            'Email': operator.email,
            'Текущая метрика': f"{operator.current_metric}%",
            'Уровень опыта': f"{operator.experience_level}%",
            'Нагрузка': f"{operator.calculate_workload()}%",
            'Всего уязвимостей': len(operator.assigned_vulnerabilities)
        })

        # Пустая строка
        data.append({})

        # Заголовки уязвимостей
        data.append({
            'ID уязвимости': 'ID',
            'Название уязвимости': 'Название',
            'Описание': 'Описание',
            'Уровень риска': 'Уровень риска',
            'Статус': 'Статус',
            'CVSS Score': 'CVSS',
            'Категория': 'Категория',
            'Правки': 'Правки'
        })

        # Данные уязвимостей
        for vuln in operator.assigned_vulnerabilities:
            data.append({
                'ID уязвимости': vuln.id,
                'Название уязвимости': vuln.title,
                'Описание': vuln.description,
                'Уровень риска': vuln.severity.upper(),
                'Статус': vuln.status,
                'CVSS Score': vuln.cvss_score,
                'Категория': vuln.category,
                'Правки': vuln.modifications
            })

        # Статистика по статусам
        status_counts = {}
        for vuln in operator.assigned_vulnerabilities:
            status_counts[vuln.status] = status_counts.get(vuln.status, 0) + 1

        # Пустая строка
        data.append({})

        # Статистика
        data.append({'ID уязвимости': 'СТАТИСТИКА:', 'Название уязвимости': '', 'Описание': '', 'Уровень риска': '',
                     'Статус': '', 'CVSS Score': '', 'Категория': '', 'Правки': ''})
        for status, count in status_counts.items():
            data.append({
                'ID уязвимости': f"Статус '{status}':",
                'Название уязвимости': count,
                'Описание': '',
                'Уровень риска': '',
                'Статус': '',
                'CVSS Score': '',
                'Категория': '',
                'Правки': ''
            })

        filename = f"уязвимости_{operator.name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return self.export_to_excel(data, filename.replace('.xlsx', ''))

    def export_performance_report(self, performance_data: dict) -> str:
        """Экспорт отчета по производительности"""
        data = []
        for op_id, metrics in performance_data.items():
            data.append({
                'ID оператора': op_id,
                'Имя оператора': metrics['name'],
                'Текущая метрика': f"{metrics['metric']}%",
                'Выполненные задачи': metrics['completed'],
                'Ожидающие задачи': metrics['pending'],
                'Нагрузка': f"{metrics['workload']}%"
            })

        return self.export_to_excel(data, "отчет_производительности")